import pandas as pd
import numpy as np
import re
from typing import List, Tuple, Optional, Union
from pathlib import Path
import openpyxl
import logging
from thefuzz import fuzz

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def flatten_headers(header_block: pd.DataFrame, sep: str = " ") -> List[str]:
    arr = header_block.astype(str).fillna("").values
    rows, cols = arr.shape
    flat = []
    for c in range(cols):
        parts = []
        for r in range(rows):
            cell = arr[r, c].strip()
            if cell and cell.lower() not in ("nan", "none"):
                parts.append(cell)
        name = sep.join(parts)
        name = re.sub(r'\s{2,}', ' ', name).strip()
        flat.append(name or f"Column_{c}")
    return flat

def clean_context(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replace blank or all‐whitespace cells with NaN, then drop any rows
    that are completely empty.
    """
    # Drop rows that are all NaN
    df_clean = df.dropna(how='all').reset_index(drop=True)
    return df_clean
class FreightTableExtractor:
    def __init__(self,ignored_sheets, custom_terms=None):
        # Terms for header scoring
        self.default_location_terms = {'origin','destination','port','pol','pod','country','area',
                               'carrier','carriers','from','to','via','start'}
        self.default_container_terms = {"20'","40'",'dc','hc','rf','rq','box','soc',
                                '20rf','40rf','20rq','40rq','dry','reefer','container'}
        self.default_rate_terms = {'rate','currency','charges','price','cost','amount','fee','tariff'}
        self.default_logistics_terms = {'mode','term','code','routing','service','transit'}
        
        # Merge with custom terms if provided
        if custom_terms:
            self.location_terms = self.default_location_terms | set(custom_terms.get('location', []))
            self.container_terms = self.default_container_terms | set(custom_terms.get('container', []))
            self.rate_terms = self.default_rate_terms | set(custom_terms.get('rate', []))
            self.logistics_terms = self.default_logistics_terms | set(custom_terms.get('logistics', []))
        else:
            self.location_terms = self.default_location_terms
            self.container_terms = self.default_container_terms
            self.rate_terms = self.default_rate_terms
            self.logistics_terms = self.default_logistics_terms
        
        self.ignored_sheets = ignored_sheets

        self.valid_header_terms = (
            self.location_terms |
            self.container_terms |
            self.rate_terms |
            self.logistics_terms
        )
        # Keywords for fuzzy matching sheet names
        self.freetime_keywords = ["free time","freetime","demurrage","detention","storage"]
        self.rule_keywords = ["rule","policy","term","condition","regulation","note","remark"]
        self.surcharges_keywords = ["surcharge","tariff","charge"]

    def normalize_sheet_name(self, name: str) -> str:
        return re.sub(r'[^a-z0-9]', '', name.lower()) if name else ''

    def fuzzy_match_any(self, text: str, choices: List[str], threshold: int=70) -> bool:
        txt = text.lower()
        return any(fuzz.partial_ratio(txt, kw) >= threshold for kw in choices)

    def load_and_unmerge(self, file_path: Union[str,Path], sheet: str) -> pd.DataFrame:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet]
        for mr in list(ws.merged_cells.ranges):
            minc,minr,maxc,maxr = mr.bounds
            val = ws.cell(minr, minc).value
            ws.unmerge_cells(str(mr))
            for r in range(minr, maxr+1):
                for c in range(minc, maxc+1):
                    ws.cell(r,c).value = val
        data = [[c.value for c in row] for row in ws.iter_rows()]
        return pd.DataFrame(data)

    def normalize_text(self, txt: str) -> str:
        if pd.isna(txt) or not isinstance(txt,str):
            return ""
        return re.sub(r'[^\w\s]', ' ', txt.lower()).strip()

    def calculate_freight_score(self, row: List[str]) -> float:
        norm = [self.normalize_text(v) for v in row if pd.notna(v)]
        txt = ' '.join(norm)
        loc = sum(t in txt for t in self.location_terms)
        cont = sum(t in txt for t in self.container_terms)
        rate = sum(t in txt for t in self.rate_terms)
        logi = sum(t in txt for t in self.logistics_terms)
        score = loc*3 + cont*2.5 + rate*2 + logi*1.5
        cats = sum(x>0 for x in (loc,cont,rate,logi))
        if cats>=2: score *= 1.5
        if cats>=3: score *= 2
        nums = sum(bool(re.search(r'\d+',v)) for v in row if isinstance(v,str))
        if nums>=3: score *= 1.3
        return score * (0.7 + 0.3 * min(1,len(row)/10))

    def calculate_contextual_score(self, df: pd.DataFrame, idx: int, w: int=3) -> float:
        scores = []
        for i in range(max(0,idx-w), min(len(df),idx+w+1)):
            row = df.iloc[i].astype(str).tolist()
            sc = self.calculate_freight_score(row)
            scores.append(sc * (2 if i==idx else 1))
        return float(np.mean(scores)) if scores else 0.0

    def detect_header_row(self, df: pd.DataFrame, thresh: float=1.5) -> Optional[int]:
        best, idx = 0.0, None
        lim = min(len(df),50)
        scores = {}
        for i in range(lim):
            row = df.iloc[i].astype(str).tolist()
            sc = self.calculate_freight_score(row)
            cs = self.calculate_contextual_score(df,i)
            final = sc*0.7 + cs*0.3
            scores[i] = final
            if final>best and final>=thresh:
                best, idx = final, i
        if idx is None:
            for i in range(lim):
                txt = ' '.join(self.normalize_text(v) for v in df.iloc[i].astype(str))
                if ('pol' in txt and 'pod' in txt) or ('carrier' in txt and 'rate' in txt):
                    if scores.get(i,0)>=0.5:
                        idx = i
                        break
        if idx is not None:
            logger.info(f"Header at row {idx} (score {best:.2f})")
        return idx

    def merge_multi_level_headers(self, df: pd.DataFrame, hrow: int, depth: int=2) -> List[str]:
        start = max(0,hrow-depth+1)
        block = df.iloc[start:hrow+1,:].reset_index(drop=True)
        return flatten_headers(block)

    def detect_table_end(self, df: pd.DataFrame, start: int, lookback: int=8,
                        pattern_tolerance: int=2, recovery_threshold: int=3,
                        min_threshold_ratio: float=0.4) -> int:
        n = len(df)
        if start >= n: return n
        
        non_null = df.notna().sum(axis=1)
        sig = df.applymap(lambda v:
                        'num' if isinstance(v,(int,float))
                        else 'txt' if isinstance(v,str) and v.strip()
                        else 'emt').agg(''.join, axis=1)
        
        init_max = non_null[start:start+lookback].max()
        threshold = max(3, init_max * min_threshold_ratio)
        
        bad, good_streak, pattern_violations = 0, 0, 0
        last_sig = None
        
        for i in range(start, n):
            cnt, s = non_null[i], sig[i]
            
            is_bad_row = False
            
            if cnt < threshold:
                is_bad_row = True
            elif last_sig and s != last_sig:
                pattern_violations += 1
                if pattern_violations > pattern_tolerance:
                    is_bad_row = True
            
            if is_bad_row:
                bad += 1
                good_streak = 0
            else:
                good_streak += 1
                pattern_violations = max(0, pattern_violations - 1)
                last_sig = s
                
                # Recovery mechanism
                if good_streak >= recovery_threshold:
                    bad = max(0, bad - 2)
            
            if bad >= lookback:
                return max(start, i - bad + 1)
        
        return n


    def clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        return df.dropna(how='all').dropna(axis=1,how='all').reset_index(drop=True)

    def extract_raw_context(self, df: pd.DataFrame, s: int, e: int) -> Optional[pd.DataFrame]:
        parts = []
        if s>0: parts.append(df.iloc[:s])
        if e<len(df): parts.append(df.iloc[e:])
        return pd.concat(parts, ignore_index=True) if parts else None
    
    def to_be_ignored(self, name: str) -> bool:
        # return self.fuzzy_match_any(name, self.ignored_sheets, threshold=70)
            return name in self.ignored_sheets

    def is_freetime_sheet(self, name: str) -> bool:
        return self.fuzzy_match_any(name, self.freetime_keywords, threshold=70)

    def is_rule_sheet(self, name: str) -> bool:
        return self.fuzzy_match_any(name, self.rule_keywords, threshold=70)
    
    def is_surcharge_sheet(self, name: str) -> bool:
        return self.fuzzy_match_any(name, self.surcharges_keywords, threshold=70)

    def get_additional_context(self, fp: Union[str,Path]) -> List[Tuple[str,pd.DataFrame]]:
        wb = openpyxl.load_workbook(fp, data_only=True)
        out = []
        for sh in wb.sheetnames:
            if self.to_be_ignored(sh):
                continue
            df = self.load_and_unmerge(fp, sh)
            if df.empty: continue
            if self.is_freetime_sheet(sh):
                hdr = f"=== FREETIME: {sh} ==="
            elif self.is_rule_sheet(sh):
                hdr = f"=== RULES/POLICY: {sh} ==="
            else:
                continue
            out.append((hdr, df))
        return out
    
    def get_additional_surcharges(self, fp: Union[str,Path]) -> List[Tuple[str,pd.DataFrame]]:
        wb = openpyxl.load_workbook(fp, data_only=True)
        out = []
        for sh in wb.sheetnames:
            if self.to_be_ignored(sh):
                continue
            df = self.load_and_unmerge(fp, sh)
            if df.empty: continue
            if self.is_surcharge_sheet(sh):
                hdr = f"=== surcharge: {sh} ==="
            else:
                continue
            out.append((hdr, df))
        return out

    def combine_context(self,
                        main: Optional[pd.DataFrame],
                        extras: List[Tuple[str,pd.DataFrame]],
                        sheet: str) -> Optional[pd.DataFrame]:
        parts = []
        if main is not None and not main.empty:
            parts.append(pd.DataFrame([[f"=== CONTEXT FROM {sheet} ==="]]))
            parts.append(main)
        for hdr, df in extras:
            parts.append(pd.DataFrame([[hdr]]))
            parts.append(df)
        return pd.concat(parts, ignore_index=True) if parts else None

    def process_excel_file(self, file_path: Union[str,Path]) -> None:
        fp = Path(file_path)
        if not fp.exists(): raise FileNotFoundError(fp)
        out_dir = fp.parent / f"{fp.stem}_processed"
        out_dir.mkdir(exist_ok=True)

        # Always gather all freetime/rule sheets up front
        extras = self.get_additional_context(fp)
        surcharges = self.get_additional_surcharges(fp)

        wb = openpyxl.load_workbook(fp, data_only=True)

        for sh in wb.sheetnames:
            if self.is_freetime_sheet(sh) or self.is_rule_sheet(sh) or self.is_surcharge_sheet(sh) or self.to_be_ignored(sh):
                continue
            df = self.load_and_unmerge(fp, sh)
            hdr = self.detect_header_row(df)
            if hdr is None:
                freight, context = None, df.copy()
            else:
                start = hdr+1
                end   = self.detect_table_end(df, start, lookback=8)
                cols  = self.merge_multi_level_headers(df, hdr)
                tbl   = df.iloc[start:end].copy()
                tbl.columns = cols[:len(tbl.columns)]
                freight = self.clean_table(tbl)
                context = self.extract_raw_context(df, start, end)

            if freight is not None and not freight.empty:
                folder = out_dir / re.sub(r'[<>:"/\\|?*]', '_', sh)
                folder.mkdir(parents=True, exist_ok=True)
                # save freight table
                output_path = Path(folder).resolve()
                output_path.mkdir(parents=True, exist_ok=True)

                # Create the full file path
                file_path = output_path / f"{output_path.name}_freight_table.xlsx"
                freight.to_excel(file_path, index=False)
                # freight.to_excel(folder / f"{folder.name}_freight_table.xlsx", index=False)
                # combine and save context
                combined = self.combine_context(context, extras, sh)
                #combine surcharges
                surcharges_combined = self.combine_context(context,surcharges,sh)

                rest = output_path / f"{output_path.name}_surcharges.xlsx"
                with pd.ExcelWriter(rest, engine='openpyxl') as w:
                    if surcharges_combined is not None and not surcharges_combined.empty:
                        # clean out blank rows
                        surcharges_combined = clean_context(surcharges_combined)
                        surcharges_combined.to_excel(w,
                                        sheet_name='rest',
                                        index=False,
                                        header=False)

                ctxf = output_path / f"{output_path.name}_context.xlsx"
                with pd.ExcelWriter(ctxf, engine='openpyxl') as w:
                    if combined is not None and not combined.empty:
                        # clean out blank rows
                        combined = clean_context(combined)
                        combined.to_excel(w,
                                        sheet_name='Context',
                                        index=False,
                                        header=False)
                    else:
                        pd.DataFrame([["No context found"]]).to_excel(w, sheet_name='Context', index=False, header=False)

        logger.info("Processing complete.")


# # Example usage:
# if __name__ == "__main__":
#     extractor = FreightTableExtractor()
#     input_excel=freightify_frontend.uploaded_file
#     extractor.process_excel_file(input_excel)
#     print("✅ Done")
